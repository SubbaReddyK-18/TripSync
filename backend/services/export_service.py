from datetime import datetime, timezone, timedelta
from io import BytesIO

IST_OFFSET = timedelta(hours=5, minutes=30)

def now_ist():
    return datetime.now(timezone.utc) + IST_OFFSET

def to_ist(dt):
    if hasattr(dt, "strftime") and hasattr(dt, "hour"):
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone(IST_OFFSET))
    return dt
from bson.objectid import ObjectId
from config.database import get_db
from fpdf import FPDF
import csv


class ExportPDF(FPDF):
    def header(self):
        if self.page_no() > 1:
            self.set_font("Helvetica", "B", 9)
            self.set_text_color(100, 100, 100)
            self.cell(0, 8, "TripSync", align="L")
            self.cell(0, 8, f"Page {self.page_no()}", align="R", new_x="LMARGIN", new_y="NEXT")
            self.line(10, 16, 200, 16)
            self.ln(4)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 7)
        self.set_text_color(150, 150, 150)
        self.cell(0, 10, f"Generated on {now_ist().strftime('%d %b %Y, %I:%M %p IST')}", align="C")

    def section_title(self, title):
        self.set_font("Helvetica", "B", 14)
        self.set_text_color(30, 60, 110)
        self.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT")
        self.set_draw_color(30, 60, 110)
        self.line(10, self.get_y(), 200, self.get_y())
        self.ln(4)

    def sub_section(self, label, value):
        self.set_font("Helvetica", "B", 10)
        self.set_text_color(60, 60, 60)
        self.cell(50, 7, label + ":")
        self.set_font("Helvetica", "", 10)
        self.set_text_color(30, 30, 30)
        self.cell(0, 7, str(value), new_x="LMARGIN", new_y="NEXT")

    def kpi_row(self, label, value):
        self.set_font("Helvetica", "B", 11)
        self.set_text_color(30, 30, 30)
        self.cell(90, 8, label)
        self.set_font("Helvetica", "", 11)
        self.set_text_color(30, 60, 110)
        self.cell(0, 8, str(value), new_x="LMARGIN", new_y="NEXT")
        self.set_draw_color(220, 220, 225)
        self.line(10, self.get_y(), 200, self.get_y())
        self.ln(1)


# ── ANALYTICS EXPORT ──

def export_analytics_pdf(analytics_data, range_key):
    buf = BytesIO()
    pdf = ExportPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=20)

    pdf.add_page()
    pdf.ln(30)
    pdf.set_font("Helvetica", "B", 26)
    pdf.set_text_color(30, 60, 110)
    pdf.cell(0, 15, "TripSync", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)
    pdf.set_font("Helvetica", "", 12)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 8, "Platform Analytics Report", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    pdf.set_font("Helvetica", "I", 10)
    range_label = range_key.title() if range_key != "all" else "All Time"
    pdf.cell(0, 6, f"Period: {range_label}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"Generated: {now_ist().strftime('%d %b %Y, %I:%M %p IST')}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(8)
    pdf.set_draw_color(30, 60, 110)
    pdf.line(60, pdf.get_y(), 150, pdf.get_y())
    pdf.add_page()

    pdf.section_title("Platform Summary")
    pdf.kpi_row("Total Users", str(analytics_data.get("hero_users", 0)))
    pdf.kpi_row("Total Admins", str(analytics_data.get("hero_admins", 0)))
    pdf.ln(4)

    pdf.section_title("Analytics KPIs")
    for key, label in [
        ("active_users", "Active Users"),
        ("trips_created", "Trips Created"),
        ("expenses_logged", "Expenses Logged"),
        ("expense_volume", "Expense Volume"),
        ("memories_uploaded", "Memories Uploaded"),
        ("places_added", "Places Added"),
        ("settlements_completed", "Settlements Completed"),
        ("new_users", "New Users"),
    ]:
        stat = analytics_data.get(key, {})
        val = stat.get("value", 0)
        growth = stat.get("growth")
        display = f"Rs. {val/100:,.2f}" if key == "expense_volume" else str(val)
        if growth is not None:
            display += f"  ({'+' if growth > 0 else ''}{growth}%)"
        pdf.kpi_row(label, display)

    pdf_bytes = pdf.output()
    buf = BytesIO(pdf_bytes)
    buf.seek(0)
    return buf


def export_analytics_csv(analytics_data, range_key):
    buf = BytesIO()
    writer = csv.writer(buf)
    writer.writerow(["Metric", "Value", "Growth (%)"])
    writer.writerow(["Period", range_key.title() if range_key != "all" else "All Time", ""])
    writer.writerow(["Total Users", analytics_data.get("hero_users", 0), ""])
    writer.writerow(["Total Admins", analytics_data.get("hero_admins", 0), ""])
    for key, label in [
        ("active_users", "Active Users"),
        ("trips_created", "Trips Created"),
        ("expenses_logged", "Expenses Logged"),
        ("expense_volume", "Expense Volume"),
        ("memories_uploaded", "Memories Uploaded"),
        ("places_added", "Places Added"),
        ("settlements_completed", "Settlements Completed"),
        ("new_users", "New Users"),
    ]:
        stat = analytics_data.get(key, {})
        val = stat.get("value", 0)
        growth = stat.get("growth", "")
        writer.writerow([label, val, growth if growth is not None else ""])
    buf.seek(0)
    return buf


def export_analytics_xlsx(analytics_data, range_key):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Analytics"
    ws.append(["TripSync - Platform Analytics"])
    ws.append(["Period", range_key.title() if range_key != "all" else "All Time"])
    ws.append([])
    ws.append(["Metric", "Value", "Growth (%)"])
    ws.append(["Total Users", analytics_data.get("hero_users", 0), ""])
    ws.append(["Total Admins", analytics_data.get("hero_admins", 0), ""])
    for key, label in [
        ("active_users", "Active Users"),
        ("trips_created", "Trips Created"),
        ("expenses_logged", "Expenses Logged"),
        ("expense_volume", "Expense Volume"),
        ("memories_uploaded", "Memories Uploaded"),
        ("places_added", "Places Added"),
        ("settlements_completed", "Settlements Completed"),
        ("new_users", "New Users"),
    ]:
        stat = analytics_data.get(key, {})
        val = stat.get("value", 0)
        growth = stat.get("growth", "")
        ws.append([label, val, growth if growth is not None else ""])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_analytics(analytics_data, range_key, fmt):
    if fmt == "pdf":
        return export_analytics_pdf(analytics_data, range_key), "application/pdf", "analytics_report.pdf"
    elif fmt == "csv":
        return export_analytics_csv(analytics_data, range_key), "text/csv", "analytics_report.csv"
    elif fmt == "xlsx":
        return export_analytics_xlsx(analytics_data, range_key), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "analytics_report.xlsx"


# ── ACTIVITY LOGS EXPORT ──

def export_logs_csv(logs):
    buf = BytesIO()
    writer = csv.writer(buf)
    writer.writerow(["Timestamp", "User", "Action", "Details", "IP Address"])
    for log in logs:
        ts = log.get("timestamp", "")
        if hasattr(ts, "strftime"):
            ts = to_ist(ts).strftime("%d %b %Y %I:%M %p")
        user = log.get("user_name", log.get("user_id", ""))
        action = log.get("action_type", "")
        details = log.get("details", "")
        ip = log.get("ip_address", "")
        writer.writerow([ts, user, action, details, ip])
    buf.seek(0)
    return buf


def export_logs_xlsx(logs):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Activity Logs"
    ws.append(["Timestamp", "User", "Action", "Details", "IP Address"])
    for log in logs:
        ts = log.get("timestamp", "")
        if hasattr(ts, "strftime"):
            ts = to_ist(ts).strftime("%d %b %Y %I:%M %p")
        user = log.get("user_name", log.get("user_id", ""))
        action = log.get("action_type", "")
        details = log.get("details", "")
        ip = log.get("ip_address", "")
        ws.append([ts, user, action, details, ip])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_logs_pdf(logs):
    buf = BytesIO()
    pdf = ExportPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(30, 60, 110)
    pdf.cell(0, 10, "TripSync - Activity Logs", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(8)

    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(240, 240, 245)
    col_w = [35, 50, 35, 100, 40]
    headers = ["Timestamp", "User", "Action", "Details", "IP Address"]
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 6, h, border=1, fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    for log in logs:
        ts = log.get("timestamp", "")
        if hasattr(ts, "strftime"):
            ts = to_ist(ts).strftime("%d %b %Y %I:%M %p")
        user = log.get("user_name", log.get("user_id", ""))
        action = log.get("action_type", "")
        details = log.get("details", "")
        ip = log.get("ip_address", "")
        if len(details) > 80:
            details = details[:77] + "..."
        pdf.cell(col_w[0], 5, str(ts)[:18], border=1)
        pdf.cell(col_w[1], 5, str(user)[:25], border=1)
        pdf.cell(col_w[2], 5, str(action)[:18], border=1)
        pdf.cell(col_w[3], 5, str(details), border=1)
        pdf.cell(col_w[4], 5, str(ip)[:15], border=1)
        pdf.ln()
        if pdf.get_y() > 185:
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_fill_color(240, 240, 245)
            for i, h in enumerate(headers):
                pdf.cell(col_w[i], 6, h, border=1, fill=True)
            pdf.ln()
            pdf.set_font("Helvetica", "", 7)

    pdf_bytes = pdf.output()
    buf = BytesIO(pdf_bytes)
    buf.seek(0)
    return buf


def export_logs(logs, fmt):
    if fmt == "csv":
        return export_logs_csv(logs), "text/csv", "activity_logs.csv"
    elif fmt == "xlsx":
        return export_logs_xlsx(logs), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "activity_logs.xlsx"
    elif fmt == "pdf":
        return export_logs_pdf(logs), "application/pdf", "activity_logs.pdf"


# ── TRIP REPORT EXPORT (multi-format) ──

def export_trip_report_xlsx(trip, members, user_map, expenses, settlements, locations, budget):
    from openpyxl import Workbook
    wb = Workbook()

    # Trip Info
    ws = wb.active
    ws.title = "Trip Info"
    ws.append(["TripSync - Trip Report"])
    ws.append(["Trip Name", trip.get("title", "Untitled")])
    ws.append(["Destination", trip.get("destination", "")])
    start = trip.get("start_date", "")
    end = trip.get("end_date", "")
    if hasattr(start, "strftime"):
        start = to_ist(start).strftime("%d %b %Y")
    if hasattr(end, "strftime"):
        end = to_ist(end).strftime("%d %b %Y")
    ws.append(["Dates", f"{start} - {end}"])
    ws.append(["Members", len(members)])

    # Members
    ws_members = wb.create_sheet("Members")
    ws_members.append(["Name", "Role", "Joined"])
    for m in members:
        uid = str(m["user_id"])
        u = user_map.get(uid, {})
        name = u.get("full_name") or u.get("username") or "Unknown"
        role = m.get("role", "viewer")
        joined = m.get("joined_at", "")
        if hasattr(joined, "strftime"):
            joined = to_ist(joined).strftime("%d %b %Y")
        ws_members.append([name, role, joined])

    # Expenses
    ws_exp = wb.create_sheet("Expenses")
    ws_exp.append(["Title", "Category", "Paid By", "Amount", "Date"])
    for e in expenses:
        paid_by = e.get("paid_by", "")
        if isinstance(paid_by, dict):
            payer_parts = []
            for pid, amt in paid_by.items():
                name = user_map.get(str(pid), {}).get("full_name") or user_map.get(str(pid), {}).get("username") or str(pid)[:8]
                payer_parts.append(f"{name} (Rs. {amt/100:.2f})")
            payer = ", ".join(payer_parts)
        else:
            pid = str(paid_by)
            payer = user_map.get(pid, {}).get("full_name") or user_map.get(pid, {}).get("username") or pid[:8]
        d = e.get("date", "")
        if hasattr(d, "strftime"):
            d = to_ist(d).strftime("%d %b %Y")
        ws_exp.append([e.get("title", ""), e.get("category", ""), payer, e.get("amount", 0) / 100, d])

    total_spent = sum(e["amount"] for e in expenses)

    # Budget
    ws_bud = wb.create_sheet("Budget")
    total_budget = budget["total_amount"] if budget else 0
    remaining = max(0, total_budget - total_spent) if total_budget else 0
    ws_bud.append(["Metric", "Value"])
    ws_bud.append(["Total Budget", f"Rs. {total_budget/100:,.2f}" if total_budget else "Not Set"])
    ws_bud.append(["Spent", f"Rs. {total_spent/100:,.2f}"])
    ws_bud.append(["Remaining", f"Rs. {remaining/100:,.2f}"])

    # Settlements
    ws_set = wb.create_sheet("Settlements")
    ws_set.append(["From", "To", "Amount", "Status", "Date"])
    for s in settlements:
        from_id = str(s.get("from_user_id", ""))
        to_id = str(s.get("to_user_id", ""))
        from_name = user_map.get(from_id, {}).get("full_name") or user_map.get(from_id, {}).get("username") or from_id[:8]
        to_name = user_map.get(to_id, {}).get("full_name") or user_map.get(to_id, {}).get("username") or to_id[:8]
        d = s.get("paid_at") or s.get("created_at", "")
        if hasattr(d, "strftime"):
            d = to_ist(d).strftime("%d %b %Y")
        ws_set.append([from_name, to_name, s.get("amount", 0) / 100, s.get("status", ""), d])

    # Places
    ws_pl = wb.create_sheet("Places")
    ws_pl.append(["Name", "Category", "Coordinates", "Added By", "Visit Date"])
    for loc in locations:
        added_by_id = str(loc.get("added_by", ""))
        added_by = user_map.get(added_by_id, {}).get("full_name") or user_map.get(added_by_id, {}).get("username") or ""
        vdate = loc.get("visit_date", "")
        if hasattr(vdate, "strftime"):
            vdate = to_ist(vdate).strftime("%d %b %Y")
        ws_pl.append([loc.get("name", ""), loc.get("category", ""), f"{loc.get('latitude', '')}, {loc.get('longitude', '')}", added_by, vdate])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_trip_report_csv(trip, members, user_map, expenses, settlements, locations, budget):
    buf = BytesIO()
    writer = csv.writer(buf)
    writer.writerow(["TripSync - Trip Report"])
    writer.writerow(["Trip", trip.get("title", "Untitled")])
    writer.writerow(["Destination", trip.get("destination", "")])
    start = trip.get("start_date", "")
    end = trip.get("end_date", "")
    if hasattr(start, "strftime"):
        start = to_ist(start).strftime("%d %b %Y")
    if hasattr(end, "strftime"):
        end = to_ist(end).strftime("%d %b %Y")
    writer.writerow(["Dates", f"{start} - {end}"])
    writer.writerow([])
    writer.writerow(["--- Members ---"])
    writer.writerow(["Name", "Role"])
    for m in members:
        uid = str(m["user_id"])
        u = user_map.get(uid, {})
        name = u.get("full_name") or u.get("username") or "Unknown"
        writer.writerow([name, m.get("role", "")])
    writer.writerow([])
    writer.writerow(["--- Expenses ---"])
    writer.writerow(["Title", "Category", "Paid By", "Amount"])
    for e in expenses:
        paid_by = e.get("paid_by", "")
        if isinstance(paid_by, dict):
            payer_parts = []
            for pid, amt in paid_by.items():
                name = user_map.get(str(pid), {}).get("full_name") or user_map.get(str(pid), {}).get("username") or str(pid)[:8]
                payer_parts.append(f"{name} (Rs. {amt/100:.2f})")
            payer = ", ".join(payer_parts)
        else:
            pid = str(paid_by)
            payer = user_map.get(pid, {}).get("full_name") or user_map.get(pid, {}).get("username") or pid[:8]
        writer.writerow([e.get("title", ""), e.get("category", ""), payer, f"Rs. {e.get('amount', 0)/100:,.2f}"])
    writer.writerow([])
    writer.writerow(["--- Settlements ---"])
    writer.writerow(["From", "To", "Amount", "Status"])
    for s in settlements:
        from_id = str(s.get("from_user_id", ""))
        to_id = str(s.get("to_user_id", ""))
        from_name = user_map.get(from_id, {}).get("full_name") or user_map.get(from_id, {}).get("username") or from_id[:8]
        to_name = user_map.get(to_id, {}).get("full_name") or user_map.get(to_id, {}).get("username") or to_id[:8]
        writer.writerow([from_name, to_name, f"Rs. {s.get('amount', 0)/100:,.2f}", s.get("status", "")])
    writer.writerow([])
    writer.writerow(["--- Places ---"])
    writer.writerow(["Name", "Category", "Visit Date"])
    for loc in locations:
        vdate = loc.get("visit_date", "")
        if hasattr(vdate, "strftime"):
            vdate = to_ist(vdate).strftime("%d %b %Y")
        writer.writerow([loc.get("name", ""), loc.get("category", ""), vdate])
    buf.seek(0)
    return buf
